import sys
import openpyxl
from googlesearch import search

def fill_urls(excel_file, wb, urls):
    # Open the Excel spreadsheet
    sheet = wb["Sheet1"]
    print(f"Opening Spreadsheet {excel_file}")
    # Loop through the rows in the spreadsheet
    for row in range(2, sheet.max_row + 1):
        # Get the company name and URL from the spreadsheet
        company_name = sheet.cell(row, 1).value
        company_url = sheet.cell(row, 2).value
        
        # If the company URL is not already filled in
        if not company_url:
            
            # Do a Google search for the company name
            # Docs: https://python-googlesearch.readthedocs.io/en/latest/
            for url in search(query=bytes(company_name, 'utf-8'), num=1, stop=1, pause=2):
                company_url = url

            print(company_name)
            print(f"Populating {excel_file} at row {row}: {company_url}")
            print('----------')

            # Update the company URL in the spreadsheet
            sheet.cell(row, 2).value = company_url
            urls.append(company_url)
    

if __name__=="__main__":
    excel_file = "companies.xlsx"

    wb = openpyxl.load_workbook(excel_file)
    urls = []
    try:
        modified_wb = fill_urls(excel_file, wb, urls)
    except Exception as e:
        print(f"An error occurred: {e}")

    if len(urls) > 0:
        # save if modified even if it fails
        print(f"Saving {len(urls)} urls to {excel_file}")
        wb.save(excel_file)
    else:
        print("No urls fetched. Excel workbook has all urls populated.")

    wb.close()
    input("Press Enter to exit...")
    sys.exit()